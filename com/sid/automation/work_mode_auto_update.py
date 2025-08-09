"""
work_mode_auto_update.py

Automatically logs your work mode into an Excel table based on the Wi-Fi network
you are connected to and the current day.

Features:
- Detects current Wi-Fi SSID on Windows
- Logs Date (DD/MM/YYYY), Day, Time, Wi-Fi SSID, and Work Type (Work From Office/Home/Weekend)
- Centers data in the Excel table cells
- Expands the Excel table as needed
- Skips logging if data for the current date already exists

Configuration:
- Update EXCEL_PATH, SHEET_NAME, and TABLE_NAME to your Excel file/table.

Author: Siddhant Patni
Date: 2025-08-09
"""

import subprocess
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# === CONFIGURATION ===
EXCEL_PATH = r"C:\Users\Siddhant Patni\Downloads\Work_Type_Tracker.xlsx"
SHEET_NAME = "Log_Data"
TABLE_NAME = "WorkLogTable"
HEADERS = ["Date", "Day", "Time", "Wi-Fi SSID", "Type of Work"]

def get_wifi_ssid() -> str | None:
    """
    Fetches the current connected Wi-Fi SSID on a Windows machine.
    Returns None if no connection or error occurs.
    """
    try:
        output = subprocess.check_output("netsh wlan show interfaces", shell=True, text=True)
        for line in output.splitlines():
            if "SSID" in line and "BSSID" not in line:
                # Format expected: 'SSID                   : YourSSID'
                return line.split(":", 1)[1].strip()
    except Exception as e:
        print(f"[!] Failed to get Wi-Fi SSID: {e}")
    return None

def determine_work_type(ssid: str | None, day_name: str) -> str:
    """
    Determines the type of work based on SSID and day of the week.
    Weekends are labeled 'Weekend'.
    SSID containing 'hilti' (case-insensitive) means 'Work From Office',
    otherwise 'Work From Home'.
    """
    if day_name in {"Saturday", "Sunday"}:
        return "Weekend"
    if ssid and "Company_Wifi_Name" in ssid.lower():
        return "Work From Office"
    return "Work From Home"

def get_table(ws, table_name: str):
    """
    Retrieves an Excel table object by name from the worksheet.
    Returns None if not found.
    """
    return next((tbl for tbl in ws.tables.values() if tbl.name == table_name), None)

def date_already_logged(ws, start_row: int, end_row: int, date_str: str) -> bool:
    """
    Checks if the given date (in DD/MM/YYYY format) already exists
    in the first column (Date) of the table.
    """
    for row in range(start_row + 1, end_row + 1):  # Skip header
        cell_value = ws.cell(row=row, column=1).value
        if cell_value == date_str:
            return True
    return False

def find_insert_row(ws, start_row: int, end_row: int, num_cols: int) -> int:
    """
    Finds the first empty row inside the table for inserting new data.
    Returns the row number. If none found, returns the next row after end_row.
    """
    for row in range(start_row + 1, end_row + 1):
        if all(ws.cell(row=row, column=col).value in (None, "") for col in range(1, num_cols + 1)):
            return row
    return end_row + 1

def append_to_excel_table(path: str, sheet_name: str, table_name: str, row_data: list[str]) -> None:
    """
    Appends a new row of data to the specified Excel table if today's date
    is not already logged. Inserts row centered and expands the table range if needed.
    """
    try:
        wb = load_workbook(path)
        ws = wb[sheet_name]

        table = get_table(ws, table_name)
        if table is None:
            print(f"[!] Table '{table_name}' not found in sheet '{sheet_name}'.")
            return

        # Parse table boundaries
        ref_start, ref_end = table.ref.split(":")
        start_col = ''.join(filter(str.isalpha, ref_start))
        start_row = int(''.join(filter(str.isdigit, ref_start)))
        end_col = ''.join(filter(str.isalpha, ref_end))
        end_row = int(''.join(filter(str.isdigit, ref_end)))

        num_cols = ord(end_col.upper()) - ord(start_col.upper()) + 1
        date_to_check = row_data[0]  # Date string in DD/MM/YYYY format

        # Skip if today's date is already logged
        if date_already_logged(ws, start_row, end_row, date_to_check):
            print(f"[!] Entry for date {date_to_check} already exists. Skipping insert.")
            return

        # Find row to insert new data
        insert_row = find_insert_row(ws, start_row, end_row, num_cols)

        # Expand table reference if inserting beyond current range
        if insert_row > end_row:
            table.ref = f"{start_col}{start_row}:{end_col}{insert_row}"

        # Insert data and apply center alignment
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=insert_row, column=col_index, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(path)
        print(f"[✓] Logged data at row {insert_row} into table '{table_name}'.")

    except Exception as e:
        print(f"[!] Error updating Excel table: {e}")

def main():
    """
    Main function to get Wi-Fi SSID, determine work type,
    and append log entry to Excel table if not already logged.
    """
    ssid = get_wifi_ssid()
    if not ssid:
        print("[!] No Wi-Fi connected. Skipping log.")
        return

    now = datetime.datetime.now()
    date_str = now.strftime("%d/%m/%Y")  # Format date as DD/MM/YYYY
    day_str = now.strftime("%A")          # Full weekday name
    time_str = now.strftime("%H:%M:%S")  # Time in 24h format

    work_type = determine_work_type(ssid, day_str)

    row = [date_str, day_str, time_str, ssid, work_type]
    append_to_excel_table(EXCEL_PATH, SHEET_NAME, TABLE_NAME, row)

if __name__ == "__main__":
    main()

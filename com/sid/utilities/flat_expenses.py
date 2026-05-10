# ==========================================================
# FLAT EXPENSE MANAGEMENT SYSTEM
# ==========================================================

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import (
    PieChart,
    BarChart,
    LineChart,
    Reference
)
from openpyxl.workbook.defined_name import DefinedName


class FlatExpenseWorkbook:

    def __init__(self):

        self.file_name = "Flat_Expense_Management.xlsx"

        self.wb = Workbook()

        # ==================================================
        # COLORS
        # ==================================================

        self.dashboard_fill = PatternFill(
            start_color="1F4E78",
            end_color="1F4E78",
            fill_type="solid"
        )

        self.expense_fill = self.dashboard_fill
        self.flat_fill = self.dashboard_fill

        self.analytics_fill = PatternFill(
            start_color="7D3C98",
            end_color="7D3C98",
            fill_type="solid"
        )

        self.master_fill = PatternFill(
            start_color="CA6F1E",
            end_color="CA6F1E",
            fill_type="solid"
        )

        self.light_fill = PatternFill(
            start_color="F8F9F9",
            end_color="F8F9F9",
            fill_type="solid"
        )

        self.yellow_fill = PatternFill(
            start_color="FFF176",
            end_color="FFF176",
            fill_type="solid"
        )

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        self.currency_format = '₹#,##0.00'
        self.date_format = 'DD-MMM-YYYY'

    # ==================================================
    # COMMON METHODS
    # ==================================================

    def style_header(self, cell, fill):

        cell.font = Font(
            color="FFFFFF",
            bold=True,
            size=11
        )

        cell.fill = fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = self.border

    def auto_adjust_columns(self, ws):

        for column_cells in ws.columns:

            max_length = 0

            column = column_cells[0].column

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except:
                    pass

            adjusted_width = min(max_length + 5, 40)

            ws.column_dimensions[
                get_column_letter(column)
            ].width = adjusted_width

    def create_table(
            self,
            ws,
            start_cell,
            end_cell,
            table_name
    ):

        table = Table(
            displayName=table_name,
            ref=f"{start_cell}:{end_cell}"
        )

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = style

        ws.add_table(table)

    def apply_borders(
            self,
            ws,
            start_row,
            end_row,
            start_col,
            end_col
    ):

        for row in ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col,
                max_col=end_col
        ):

            for cell in row:
                cell.border = self.border

    # ==================================================
    # DYNAMIC NAMED RANGES
    # ==================================================

    def create_dynamic_named_ranges(self):

        self.wb.defined_names.add(
            DefinedName(
                "ExpenseCategories",
                attr_text=(
                    "'Master Data'!$C$2:"
                    "INDEX('Master Data'!$C:$C,"
                    "COUNTA('Master Data'!$C:$C))"
                )
            )
        )

        self.wb.defined_names.add(
            DefinedName(
                "PaymentModes",
                attr_text=(
                    "'Master Data'!$D$2:"
                    "INDEX('Master Data'!$D:$D,"
                    "COUNTA('Master Data'!$D:$D))"
                )
            )
        )

    # ==================================================
    # MASTER DATA
    # ==================================================

    def create_master_data_sheet(self):

        ws = self.wb.create_sheet("Master Data")

        ws.sheet_view.showGridLines = False

        headers = [
            "Years",
            "Months",
            "Expense Categories",
            "Payment Modes"
        ]

        for idx, header in enumerate(headers, start=1):

            cell = ws.cell(row=1, column=idx)

            cell.value = header

            self.style_header(
                cell,
                self.master_fill
            )

        years = list(range(2024, 2061))

        months = [
            "January", "February", "March",
            "April", "May", "June",
            "July", "August", "September",
            "October", "November", "December"
        ]

        categories = sorted([
            "Appliances",
            "Cleaning",
            "Electricity",
            "Furniture",
            "Gas",
            "Insurance",
            "Internet",
            "Maintenance",
            "Miscellaneous",
            "Property Tax",
            "Repairs",
            "Society Charges",
            "Water Bill",
        ])

        payment_modes = [
            "Cash",
            "UPI",
            "Credit Card",
            "Debit Card",
            "Bank Transfer",
            "Cheque",
            "Net Banking"
        ]

        for idx, value in enumerate(years, start=2):
            ws[f"A{idx}"] = value

        for idx, value in enumerate(months, start=2):
            ws[f"B{idx}"] = value

        for idx, value in enumerate(categories, start=2):
            ws[f"C{idx}"] = value

        for idx, value in enumerate(payment_modes, start=2):
            ws[f"D{idx}"] = value

        ws.freeze_panes = "A2"

        self.apply_borders(
            ws,
            1,
            40,
            1,
            4
        )

        self.auto_adjust_columns(ws)

    # ==================================================
    # DASHBOARD
    # ==================================================

    def create_dashboard_sheet(self):

        ws = self.wb.active

        ws.title = "Dashboard"

        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:J2")

        ws["A1"] = "FLAT EXPENSE MANAGEMENT DASHBOARD"

        ws["A1"].font = Font(
            size=22,
            bold=True,
            color="FFFFFF"
        )

        ws["A1"].fill = self.dashboard_fill

        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # ==================================================
        # PROPERTY INFORMATION
        # ==================================================

        ws["A4"] = "Property Information"

        self.style_header(
            ws["A4"],
            self.dashboard_fill
        )

        property_fields = [
            ("Flat Number", "='Flat Details'!B4"),
            ("Building Name", "='Flat Details'!B5"),
            ("Owner Name", "='Flat Details'!B7"),
            ("City", "='Flat Details'!B11"),
            ("Property Size", "='Flat Details'!B16"),
        ]

        row_num = 5

        for field, formula in property_fields:

            ws[f"A{row_num}"] = field
            ws[f"B{row_num}"] = formula

            ws[f"A{row_num}"].font = Font(bold=True)

            row_num += 1

        # ==================================================
        # KPI SECTION
        # ==================================================

        kpis = [
            (
                "Total Expenses",
                "=SUM('Flat Expenses'!J2:J1000)"
            ),
            (
                "Highest Expense",
                "=MAX('Flat Expenses'!J2:J1000)"
            ),
            (
                "Lowest Expense",
                "=MIN('Flat Expenses'!J2:J1000)"
            ),
            (
                "Total Transactions",
                '=COUNTIF(\'Flat Expenses\'!J2:J1000,">0")'
            ),
        ]

        start_row = 5

        for idx, kpi in enumerate(kpis):

            col = 4 + (idx * 2)

            ws.cell(
                row=start_row,
                column=col
            ).value = kpi[0]

            self.style_header(
                ws.cell(
                    row=start_row,
                    column=col
                ),
                self.dashboard_fill
            )

            ws.cell(
                row=start_row + 1,
                column=col
            ).value = kpi[1]

            ws.cell(
                row=start_row + 1,
                column=col
            ).fill = self.light_fill

            ws.cell(
                row=start_row + 1,
                column=col
            ).border = self.border

        # ==================================================
        # CATEGORY SUMMARY
        # ==================================================

        ws["A14"] = "Expense Category"
        ws["B14"] = "Total Amount"

        self.style_header(ws["A14"], self.dashboard_fill)
        self.style_header(ws["B14"], self.dashboard_fill)

        for row in range(2, 32):

            excel_row = row + 13

            ws[f"A{excel_row}"] = (
                f"='Master Data'!C{row}"
            )

            ws[f"B{excel_row}"] = (
                f'=IF(A{excel_row}="", "", '
                f'SUMIF('
                f'\'Flat Expenses\'!E:E,'
                f'A{excel_row},'
                f'\'Flat Expenses\'!J:J))'
            )

            ws[f"B{excel_row}"].number_format = (
                self.currency_format
            )

        self.apply_borders(
            ws,
            14,
            45,
            1,
            2
        )

        # ==================================================
        # PIE CHART
        # ==================================================

        pie = PieChart()

        labels = Reference(
            ws,
            min_col=1,
            min_row=15,
            max_row=37
        )

        data = Reference(
            ws,
            min_col=2,
            min_row=14,
            max_row=37
        )

        pie.add_data(
            data,
            titles_from_data=True
        )

        pie.set_categories(labels)

        pie.title = "Expense Distribution"

        pie.width = 14
        pie.height = 10

        ws.add_chart(pie, "D14")

        # ==================================================
        # UPDATED ON SECTION
        # ==================================================

        ws.merge_cells("A48:D49")

        ws["A48"] = '=CONCAT("Last Updated On : ",TEXT(NOW(),"DD-MMM-YYYY HH:MM:SS"))'

        ws["A48"].font = Font(
            bold=True,
            color="FFFFFF",
            size=11
        )

        ws["A48"].fill = self.dashboard_fill

        ws["A48"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        ws["A48"].border = self.border

        self.auto_adjust_columns(ws)

    # ==================================================
    # FLAT DETAILS
    # ==================================================

    def create_flat_details_sheet(self):

        ws = self.wb.create_sheet("Flat Details")

        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:D2")

        ws["A1"] = "FLAT DETAILS"

        ws["A1"].font = Font(
            size=18,
            bold=True,
            color="FFFFFF"
        )

        ws["A1"].fill = self.flat_fill

        fields = [
            "Flat Number",
            "Building Name",
            "Society Name",
            "Owner Name",
            "Contact Number",
            "Email",
            "Flat Address",
            "City",
            "State",
            "PIN Code",
            "Google Maps Link",
            "Purchase Date",
            "Property Size (sq ft)",
            "Parking Details",
            "Notes"
        ]

        for idx, field in enumerate(fields, start=4):

            ws[f"A{idx}"] = field

            self.style_header(
                ws[f"A{idx}"],
                self.flat_fill
            )

            ws[f"B{idx}"].border = self.border

        self.apply_borders(
            ws,
            4,
            18,
            1,
            2
        )

        self.auto_adjust_columns(ws)

    # ==================================================
    # FLAT EXPENSES
    # ==================================================

    def create_flat_expenses_sheet(self):

        ws = self.wb.create_sheet("Flat Expenses")

        ws.sheet_view.showGridLines = False

        columns = [
            "Sr No.",
            "Date",
            "Year",
            "Month",
            "Expense Category",
            "Description",
            "Vendor / Paid To",
            "Payment Mode",
            "Transaction Reference",
            "Amount",
            "Remarks",
            "Filtered Total"
        ]

        for idx, column in enumerate(columns, start=1):

            cell = ws.cell(row=1, column=idx)

            cell.value = column

            self.style_header(
                cell,
                self.expense_fill
            )

        # ==================================================
        # FILTERED TOTAL
        # ==================================================

        ws["L2"] = "=SUBTOTAL(9,J2:J1000)"

        ws["L2"].font = Font(
            bold=True,
            color="000000",
            size=12
        )

        ws["L2"].fill = self.yellow_fill

        ws["L2"].number_format = self.currency_format

        ws["L2"].border = self.border

        for row in range(2, 1001):

            ws[f"A{row}"] = row - 1

            ws[f"B{row}"].number_format = (
                self.date_format
            )

            ws[f"C{row}"] = (
                f'=IF(B{row}="","",YEAR(B{row}))'
            )

            ws[f"D{row}"] = (
                f'=IF(B{row}="","",TEXT(B{row},"MMMM"))'
            )

            ws[f"J{row}"].number_format = (
                self.currency_format
            )

            if row % 2 == 0:

                for col in range(1, 13):

                    ws.cell(
                        row=row,
                        column=col
                    ).fill = self.light_fill

        category_validation = DataValidation(
            type="list",
            formula1="=ExpenseCategories"
        )

        payment_validation = DataValidation(
            type="list",
            formula1="=PaymentModes"
        )

        ws.add_data_validation(category_validation)
        ws.add_data_validation(payment_validation)

        category_validation.add("E2:E1000")
        payment_validation.add("H2:H1000")

        self.create_table(
            ws,
            "A1",
            "L1000",
            "FlatExpenseTable"
        )

        self.apply_borders(
            ws,
            1,
            1000,
            1,
            12
        )

        ws.freeze_panes = "A2"

        ws.column_dimensions["L"].width = 20

        self.auto_adjust_columns(ws)

    # ==================================================
    # ANALYTICS
    # ==================================================

    def create_analytics_sheet(self):

        ws = self.wb.create_sheet("Analytics")

        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:J2")

        ws["A1"] = "EXPENSE ANALYTICS"

        ws["A1"].font = Font(
            size=22,
            bold=True,
            color="FFFFFF"
        )

        ws["A1"].fill = self.analytics_fill

        filters = [
            "Year Filter",
            "Month Filter",
            "Category Filter"
        ]

        for idx, value in enumerate(filters, start=1):

            cell = ws.cell(
                row=4,
                column=idx
            )

            cell.value = value

            self.style_header(
                cell,
                self.analytics_fill
            )

        ws["A7"] = "Expense Category"
        ws["B7"] = "Total Amount"

        self.style_header(ws["A7"], self.analytics_fill)
        self.style_header(ws["B7"], self.analytics_fill)

        for row in range(2, 32):

            excel_row = row + 6

            ws[f"A{excel_row}"] = (
                f"='Master Data'!C{row}"
            )

            ws[f"B{excel_row}"] = (
                f'=IF(A{excel_row}="", "", '
                f'SUMIF('
                f'\'Flat Expenses\'!E:E,'
                f'A{excel_row},'
                f'\'Flat Expenses\'!J:J))'
            )

        ws["D7"] = "Month"
        ws["E7"] = "Total Amount"

        self.style_header(ws["D7"], self.analytics_fill)
        self.style_header(ws["E7"], self.analytics_fill)

        for row in range(2, 14):

            excel_row = row + 6

            ws[f"D{excel_row}"] = (
                f"='Master Data'!B{row}"
            )

            ws[f"E{excel_row}"] = (
                f'=SUMIF('
                f'\'Flat Expenses\'!D:D,'
                f'D{excel_row},'
                f'\'Flat Expenses\'!J:J)'
            )

        ws["G7"] = "Year"
        ws["H7"] = "Total Amount"

        self.style_header(ws["G7"], self.analytics_fill)
        self.style_header(ws["H7"], self.analytics_fill)

        for row in range(2, 40):

            excel_row = row + 6

            ws[f"G{excel_row}"] = (
                f"='Master Data'!A{row}"
            )

            ws[f"H{excel_row}"] = (
                f'=SUMIF('
                f'\'Flat Expenses\'!C:C,'
                f'G{excel_row},'
                f'\'Flat Expenses\'!J:J)'
            )

        ws.auto_filter.ref = "A7:H50"

        self.apply_borders(
            ws,
            7,
            50,
            1,
            8
        )

        # PIE CHART

        pie = PieChart()

        labels = Reference(
            ws,
            min_col=1,
            min_row=8,
            max_row=37
        )

        data = Reference(
            ws,
            min_col=2,
            min_row=7,
            max_row=37
        )

        pie.add_data(
            data,
            titles_from_data=True
        )

        pie.set_categories(labels)

        pie.title = "Category Wise Expense Distribution"

        pie.width = 18
        pie.height = 12

        ws.add_chart(pie, "A55")

        # BAR CHART

        bar = BarChart()

        bar_data = Reference(
            ws,
            min_col=5,
            min_row=7,
            max_row=19
        )

        bar_labels = Reference(
            ws,
            min_col=4,
            min_row=8,
            max_row=19
        )

        bar.add_data(
            bar_data,
            titles_from_data=True
        )

        bar.set_categories(bar_labels)

        bar.title = "Monthly Expense Analysis"

        bar.width = 18
        bar.height = 12

        ws.add_chart(bar, "J55")

        # LINE CHART

        line = LineChart()

        line_data = Reference(
            ws,
            min_col=8,
            min_row=7,
            max_row=45
        )

        line_labels = Reference(
            ws,
            min_col=7,
            min_row=8,
            max_row=45
        )

        line.add_data(
            line_data,
            titles_from_data=True
        )

        line.set_categories(line_labels)

        line.title = "Year Wise Expense Trend"

        line.width = 28
        line.height = 14

        ws.add_chart(line, "A82")

        self.auto_adjust_columns(ws)

    # ==================================================
    # GENERATE WORKBOOK
    # ==================================================

    def generate_workbook(self):

        self.create_dashboard_sheet()

        self.create_flat_details_sheet()

        self.create_flat_expenses_sheet()

        self.create_master_data_sheet()

        self.create_dynamic_named_ranges()

        self.create_analytics_sheet()

        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(
                self.wb["Sheet"]
            )

        self.wb.save(self.file_name)

        print(
            f"Workbook generated successfully : "
            f"{self.file_name}"
        )


# ==========================================================
# MAIN
# ==========================================================

if __name__ == "__main__":

    workbook = FlatExpenseWorkbook()

    workbook.generate_workbook()